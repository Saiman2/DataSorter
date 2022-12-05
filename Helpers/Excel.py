import openpyxl


class Excel:
    sh = False
    positions = {
        'name': False,
        'price': False,
        'price_tx': False,
    }
    positionsSet = False

    def __init__(self, pathToFile=False):
        if not pathToFile:
            raise Exception("Path to file excel is required")
        # load excel with its path
        wrkbk = openpyxl.load_workbook(pathToFile)

        self.sh = wrkbk.active

    def setPosition(self, key=False, value=False):
        if not key or not value:
            raise Exception("Key & value reuqired")
        self.positions[key] = value

    def checkIfPositionsSet(self):
        if not self.positionsSet:
            set = True
            for key in self.positions:
                if self.positions[key] == False:
                    set = False
            self.positionsSet = set
        return self.positionsSet

    def foreach(self):
        sh = self.sh
        for i in range(1, sh.max_row + 1):
            print("\n")
            print("Row ", i, " data :")

            for j in range(1, sh.max_column + 1):
                cell_obj = sh.cell(row=i, column=j)
                print(cell_obj.value, end=" ")
